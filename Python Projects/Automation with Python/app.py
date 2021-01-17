import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    workbook = xl.load_workbook(filename)
    sheet = workbook['Sheet1']

    user_input_percent_to_change = float(input("Enter the percentage to change: "))
    percent_to_change = 1+user_input_percent_to_change

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        adjusted_price = cell.value * percent_to_change
        adjusted_price_cell = sheet.cell(row, 4)
        adjusted_price_cell.value = adjusted_price

        values = Reference(sheet,
                           min_row=2,
                           max_row=4,
                           min_col=4,
                           max_col=4)


    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    print("1. Save as new file")
    print("2. Save file by overriding existing file")
    user_choice = int(input("> "))
    if user_choice == 1:
        filename = str(input("Enter the file name to save as: ") + ".xlsx")
        workbook.save(filename)
        print("File has been saved as " + str(filename))
    elif user_choice == 2:
        workbook.save(filename)
        print("File has been saved (overridden the initial file)")


file_to_proceed = str(input("Enter the file name: ")+".xlsx")
process_workbook(file_to_proceed)
